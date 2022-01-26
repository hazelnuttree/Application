# $ 프로그램 개요
# 1. 주요기능 : 동일한 포맷의 엑셀 파일에 있는 값을 한 파일로 자동으로 취합하는 프로그램
# 2. 작성자 : 장홍진 팀장, 김현진 차장
# 3. 최초 개발일자 : 2022.01.12
# 4. 최초 버전 : excel_sum_v1.0  
# 5. 마지막 버전 : EXLEL_GATHER_v1.0 (2022.01.19)

# 5. 프로그램 구현 조건
# 5.1 작업대상이 되는 엑셀파일이 한 폴더에 있어야 함
# 5.2 input 폴더 내 엑셀파일은 작업대상 파일만 있어야 함
# 5.3 작업파일의 작업시트는 매 처음에 위치해야 함(첫번째 시트에 대해서만 작업)
# 5.4 작업대상이 되는 엑셀파일의 데이터 입력값은 맨 처음의 셀값에 빈값이 없어야 함
# 5.5 작업대상이 되는 엑셀파일(output file 포함)이 열려있으면 안됨

# 6. 프로그램 주요 특징
# 6.1. 사용자 입력 사항 : 
#    1) 입력폴더, 
#    2) 데이터 입력 처음 셀 주소, 
#    3) 데이터 입력 마지막 셀 주소, 
# 6.2 output 폴더 및 output 파일 자동생성

# 주의사항 : 엑셀 파일이 열려있으면 에러 발생 (input file, output file)

# $ 주요 업데이트
# === excel_sum_v1.1 ===
# 사용자 인터페이스 제작
# 취합파일 업데이트 에러 해결

# === excel_sum_v1.2 ===
# 칼럼제목 추가
# 파일명 정보를 알수 있는 칼럼 추가
# 엑셀 97버전 변환 코드 추가 : 실패

# === excel_sum_v2.2 ===
# 입력 형태 변경 (3가지 정보 입력)
# 취합.xlsx 파일이 열려있는 경우 에러 메세지 생성
# 입력값 통제

# === excel_sum_v2.3 ===> EXLEL_GATHER_v1.0
# UI 디자인 개선
# 도움말 파일 제작

# $ 향후 과제
# 열려있는 파일을 강제로 닫을 수 있는가?
# xls 파일을 대상으로 작업할 수 있는가?
# 한글명이 포함된 폴더에서의 안정성을 담보할 수 있는가?

# *********************************************
# $$ 가. 프로그램 환경설정
# ---------------------------------------------

# $$$ 1. 주요모듈
#-*- coding: utf-8 -*-

import sys
import os
import random

import datetime
import time

import pathlib

import openpyxl

import tkinter as tk
from tkinter import Label
from tkinter import Button
from tkinter import Entry
from tkinter import filedialog
from tkinter import messagebox

import PIL

# *********************************************
# UI 설정
# ---------------------------------------------

# UI 패널 객체 생성
root = tk.Tk()
root.title("EXCEL GATHER v1.0")
root.geometry("430x250+200+200")
root.resizable(False, False)

global input_folder, inputbox_folder, start_cell, end_cell, start_row_num

# 변수 초기화
inputbox_folder = ''

# 프로그램 시작
def OK():
    global input_folder, start_cell, end_cell, start_row_num

     # 프로그램 유효기간 체크
    if datetime.datetime.today() > datetime.datetime(2022,12,31) :
        tk.messagebox.showinfo("경고","프로그램 유효기간이 지났습니다. 개발자에게 문의해 주세요.")
        sys.exit()

    input_folder = inputbox_folder
    start_cell = inputbox_start.get()
    end_cell = inputbox_end.get()
    start_row_num = inputbox_row_num.get()
    
    # 입력값 유효 점검
    swich = 0 # 프로그램 실행을 위한 변수 설정 (swich = 1 일 때 실행 보류)
    
    if input_folder == '':
        tk.messagebox.showinfo("경고","취합 대상 엑셀파일이 모여 있는 폴더를 지정해 주세요.")
        swich = 1
    
    if start_cell.isalpha() == False or end_cell.isalpha() == False:
        tk.messagebox.showinfo("경고","처음 열 또는 마지막 열은 알파벳만 입력가능합니다.")
        swich = 1

    elif start_cell >= end_cell :
        tk.messagebox.showinfo("경고","마지막 열의 값은 처음 열의 값보다 커야합니다.")
        swich = 1

    if start_row_num.isdigit() == False :
        tk.messagebox.showinfo("경고","처음 행은 숫자만 입력 가능합니다.")
        swich = 1

    elif int(start_row_num) <= 0  :
        tk.messagebox.showinfo("경고","처음 행에 1 이상의 숫자를 입력하세요")
        swich = 1
        
    # UI 패널 지우기
    if swich == 0:
        root.destroy()
    
# 폴더선택
def folder():
    global inputbox_folder

    inputbox_folder = tk.filedialog.askdirectory(parent=root,initialdir="/",title='Please select a directory')    

    # 선택한 경로 보여주
    lbl1_1.configure(text="경로 : " + inputbox_folder)

    return inputbox_folder

# 도움말 팝업

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")       
    return os.path.join(base_path, relative_path)    

def help():
    image_source = resource_path("EXCEL_GATHER_HELP.png")
    help_image = PIL.Image.open(image_source)
    help_image.show()

# UI 패널 레이아웃 라인설정
line1 = 20
line2 = 50
line3 = 70
line4 = 100
line5 = 130
line6 = 160
line7 = 190
line8 = 210

# UI 패널 레이아웃 
# UI 패널 폴더선택
lbl1 = Label(root, text = " 취합할 자료들이 있는 폴더를 선택해 주세요. ", font = "NanumGothic 8")
lbl1.place(x=10, y=line1)

btn1 = Button(root, text ="폴더선택", command = folder, width=8, height=1)
btn1.place(x=260, y=line1)

# UI 패널 선택한 폴더 경로
lbl1_1 = Label(root, text="")
lbl1_1.place(x=10, y=line2)

# UI 패널 셀 선택
lbl2 = Label(root, text = " 취합 대상 자료 중 첫번째 자료의 처음 열, 마지막 열, 첫 행을 입력해 주세요.", font = "NanumGothic 8")
lbl2.place(x=10, y=line4)

lbl3 = Label(root, text = " 처음 열  ", font = "NanumGothic 8")
lbl3.place(x=10, y=line5)

inputbox_start = Entry(root, width = 5)
inputbox_start.place(x=70, y=line5)

lbl4 = Label(root, text = " 마지막 열  ", font = "NanumGothic 8")
lbl4.place(x=120, y=line5)

inputbox_end = Entry(root, width = 5)
inputbox_end.place(x=190, y=line5)

lbl5 = Label(root, text = " 처음 행  ", font = "NanumGothic 8")
lbl5.place(x=240, y=line5)

inputbox_row_num = Entry(root, width = 5)
inputbox_row_num.place(x=295, y=line5)

btn2 = Button(root, text ="OK", command = OK, width=3, height=1, foreground="blue")
btn2.place(x=350, y=line5)

# UI 패널 도움말
lbl6 = Label(root, text = "♧  I LOVE KODIT  ♧", font = "NanumGothic 8")
lbl6.place(x=10, y=line7)

lbl7 = Label(root, text = " ※ made by  장홍진 & 김현진 ", font = "NanumGothic 8")
lbl7.place(x=170, y=line8)

btn3 = Button(root, text = "도움말", command = help, width=6, height=1, relief = 'groove')
btn3.place(x=330, y=line8)

root.mainloop()

# ---------------------------------------------

print("setting")

# $$$ 2. 작업폴더 설정
rawdata_dir = pathlib.Path(input_folder)

# output 폴더 (자동생성)
output_folder = input_folder + '\\output\\'

if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# 기존 취합 xlsx 삭제
if os.path.exists(output_folder+"취합.xlsx"):
    try:
        os.remove(output_folder+"취합.xlsx")        
    except PermissionError:
        root = tk.Tk()
        root.withdraw()
        tk.messagebox.showinfo("경고", "작업 폴더 내 모든 열려있는 엑셀파일을 모두 종료 후 다시 실행하세요")
        # sys.exit()
        os._exit(00)

# ---------------------------------------------

# $$$ 3. 작업대상 파일 지정

# input 파일
input_files = rawdata_dir.glob('*.xlsx')    

# output 파일 ("취합.xlsx" 파일 자동생성)
output_wb = openpyxl.Workbook()
output_wb.active.title = "취합"
output_file = output_folder + "취합" + ".xlsx"
output_wb.save(output_file)
output_wb = openpyxl.load_workbook(output_file)
output_ws = output_wb["취합"]

# ---------------------------------------------
# $$$ 4. 작업범위 설정
# 데이터가 입력되는 첫번 째 줄의 처음과 마지막 셀 주소를 입

# 처음과 마지막 셀 주소 변수 처리

# 셀 주소의 칼럼 숫자
start_column = output_ws[start_cell+start_row_num].column
end_column = output_ws[end_cell+start_row_num].column

# 행 숫자
start_row = int(start_row_num)

# *********************************************
# $$ 나. 프로그램 구현
# ---------------------------------------------

print("start")

# output file 행번호 초기화
k = 4

# $$$ 1.input 폴더 내 input file 읽어오기
for excel_file in input_files:

    input_wb = openpyxl.load_workbook(excel_file)
    input_wb.close()
    sheet_name = input_wb.sheetnames[0]
    input_ws = input_wb[sheet_name]

    path_name = str(excel_file)
    file_name = path_name.split('\\')[-1]
    print('file_name: ', file_name)

    # input file 행번호 초기화
    i = start_row
    
    # input file 처리 건수 초기화
    cnt = 0
    
    temp1=''
    temp2=''
    
    # $$$ 2-1. 칼럼 제목 설정
    for j in range(start_column, end_column+1):
    
        # 제목복사
        if i > 2 :
            temp1 = input_ws.cell(i-2,j).value    
            output_ws.cell(2,j,value = temp1)
        if i > 1 :
            temp2 = input_ws.cell(i-1,j).value    
            output_ws.cell(3,j,value = temp2)
            output_ws.cell(3,j+1,value = "엑셀파일명")
    
    # $$$ 2-2. input file 처리(행별로 작업)
    while True:

        # $$$ 3. input file 내 셀 값 읽어오기 (열별로 작업)
        temp=''
        for j in range(start_column, end_column+1):
            
            # input file 값 임시저장
            temp = input_ws.cell(i,j).value    
            print(f'행: {i}, 열: {j}, 셀값: {input_ws.cell(i,j).value}')            
            
            # output file에 복사
            output_ws.cell(k,j,value = temp)
        
        # 맨 마지막 칼럼에 파일명 정보 추가
        output_ws.cell(k,j+1,value =file_name)
                       
        i = i+1
        cnt = cnt + 1
        k = k+1

        # input file 입력 값이 없으면 읽기 중단 
        # 각 행별로 맨 처음의 셀값을 기준으로 판단
        if input_ws.cell(i,start_column).value is None:    
            break       

print("end")

# $$$ 4. ouput file 셀 넓이 자동맞춤

for i, column_cells in enumerate(output_ws.columns):
    length = max(len(str(cell.value)) for cell in column_cells)
    output_ws.column_dimensions[column_cells[0].column_letter].width = length + 10

root = tk.Tk()
root.withdraw()
tk.messagebox.showinfo("정보", "프로그램이 " + str(cnt) + "개의 자료에서 " + str(k-4) + "건을 취합하었습니다.")
root.quit()
    
# $$$ 5. ouput file 저장
output_wb.save(output_file)
output_wb.close()    
print("success!!")

# restart kernel
os._exit(00)